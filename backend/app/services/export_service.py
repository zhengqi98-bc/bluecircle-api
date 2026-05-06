"""Data export service — CSV, Excel, PDF generation."""

import csv
import io
import os
from datetime import datetime, timezone
from typing import Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.turtle import Turtle
from app.models.track_point import TrackPoint
from app.models.alert import Alert


# ── CSV Export ───────────────────────────────────────────────────────

async def export_tracks_csv(
    db: AsyncSession,
    turtle_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
) -> str:
    """Export track points as CSV string."""
    q = select(TrackPoint).order_by(TrackPoint.time.desc())

    if turtle_id:
        q = q.where(TrackPoint.turtle_id == turtle_id)
    if from_date:
        q = q.where(TrackPoint.time >= datetime.fromisoformat(from_date))
    if to_date:
        q = q.where(TrackPoint.time <= datetime.fromisoformat(to_date))

    q = q.limit(10000)
    result = await db.execute(q)
    points = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "turtle_id", "time", "lat", "lng",
        "battery_pct", "speed_kmh", "depth_m", "temperature_c", "source"
    ])

    for p in points:
        writer.writerow([
            p.turtle_id,
            p.time.isoformat() if p.time else "",
            float(p.lat) if p.lat else "",
            float(p.lng) if p.lng else "",
            float(p.battery_pct) if p.battery_pct else "",
            float(p.speed_kmh) if p.speed_kmh else "",
            float(p.depth_m) if p.depth_m else "",
            float(p.temperature_c) if p.temperature_c else "",
            p.source or "",
        ])

    return output.getvalue()


async def export_turtles_csv(db: AsyncSession) -> str:
    """Export turtle registry as CSV."""
    result = await db.execute(select(Turtle).order_by(Turtle.id))
    turtles = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "id", "name", "name_en", "species", "species_en",
        "sex", "origin", "last_lat", "last_lng",
        "last_battery_pct", "last_speed_kmh", "last_depth_m",
        "is_active", "risk_level", "last_seen_at"
    ])

    for t in turtles:
        writer.writerow([
            t.id, t.name, t.name_en or "", t.species, t.species_en or "",
            t.sex or "", t.origin or "",
            float(t.last_lat) if t.last_lat else "",
            float(t.last_lng) if t.last_lng else "",
            float(t.last_battery_pct) if t.last_battery_pct else "",
            float(t.last_speed_kmh) if t.last_speed_kmh else "",
            float(t.last_depth_m) if t.last_depth_m else "",
            "Yes" if t.is_active else "No",
            t.risk_level or "",
            t.last_seen_at.isoformat() if t.last_seen_at else "",
        ])

    return output.getvalue()


async def export_alerts_csv(db: AsyncSession) -> str:
    """Export alerts as CSV."""
    result = await db.execute(
        select(Alert).order_by(Alert.created_at.desc()).limit(5000)
    )
    alerts = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "id", "turtle_id", "alert_type", "severity",
        "title", "description", "lat", "lng",
        "status", "created_at", "resolved_at"
    ])

    for a in alerts:
        writer.writerow([
            a.id, a.turtle_id or "", a.alert_type, a.severity,
            a.title, a.description or "",
            float(a.lat) if a.lat else "", float(a.lng) if a.lng else "",
            a.status,
            a.created_at.isoformat() if a.created_at else "",
            a.resolved_at.isoformat() if a.resolved_at else "",
        ])

    return output.getvalue()


# ── Excel Export ─────────────────────────────────────────────────────

async def export_tracks_excel(
    db: AsyncSession,
    turtle_id: Optional[str] = None,
) -> io.BytesIO:
    """Export track points as formatted Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    q = select(TrackPoint).order_by(TrackPoint.time.desc())
    if turtle_id:
        q = q.where(TrackPoint.turtle_id == turtle_id)
    q = q.limit(5000)

    result = await db.execute(q)
    points = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Track Points"

    # Header style
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1CA7A8", end_color="1CA7A8", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["海龟ID", "时间", "纬度", "经度", "电量%", "速度km/h", "深度m", "水温°C", "来源"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 24

    for row_idx, p in enumerate(points, 2):
        values = [
            p.turtle_id,
            p.time.strftime("%Y-%m-%d %H:%M") if p.time else "",
            round(float(p.lat), 4) if p.lat else "",
            round(float(p.lng), 4) if p.lng else "",
            round(float(p.battery_pct), 1) if p.battery_pct else "",
            round(float(p.speed_kmh), 2) if p.speed_kmh else "",
            round(float(p.depth_m), 1) if p.depth_m else "",
            round(float(p.temperature_c), 1) if p.temperature_c else "",
            p.source or "—",
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.font = Font(name="Arial", size=10)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── PDF Export ───────────────────────────────────────────────────────

async def export_report_pdf(html_content: str) -> io.BytesIO:
    """Convert HTML report to PDF using fpdf2.

    NOTE: fpdf2 has limited HTML support. For full HTML→PDF conversion,
    consider using weasyprint (requires system libraries) or wkhtmltopdf.
    This is a simplified PDF wrapper.
    """
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Add Unicode font (DejaVu bundled with fpdf2)
    pdf.add_font("DejaVu", "", "/opt/bluecircle-backend/venv/lib/python3.9/site-packages/fpdf/fonts/DejaVuSans.ttf", uni=True)
    pdf.add_font("DejaVu", "B", "/opt/bluecircle-backend/venv/lib/python3.9/site-packages/fpdf/fonts/DejaVuSans-Bold.ttf", uni=True)

    # Title
    pdf.set_font("DejaVu", "B", 16)
    pdf.cell(0, 10, "Blue Circle — Report", ln=True, align="C")
    pdf.set_font("DejaVu", "", 10)
    pdf.cell(0, 8, f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", ln=True, align="C")
    pdf.ln(10)

    # Extract text from HTML (simple approach — strip tags)
    import re
    text = re.sub(r"<style[^>]*>.*?</style>", "", html_content, flags=re.DOTALL)
    text = re.sub(r"<script[^>]*>.*?</script>", "", text, flags=re.DOTALL)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    pdf.set_font("DejaVu", "", 9)
    pdf.multi_cell(0, 5, text[:5000])  # Limit to avoid huge PDFs

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return output


# ── Export dispatcher ────────────────────────────────────────────────

EXPORT_FORMATS = {
    "csv": ("text/csv", ".csv"),
    "excel": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx"),
    "pdf": ("application/pdf", ".pdf"),
}

EXPORT_TYPES = ["tracks", "turtles", "alerts", "report"]


async def get_export(
    db: AsyncSession,
    export_type: str,
    format: str,
    turtle_id: Optional[str] = None,
    report_html: Optional[str] = None,
) -> tuple[bytes, str, str]:
    """Get export data in requested format.

    Returns:
        tuple: (content_bytes, content_type, filename)
    """
    content_type, ext = EXPORT_FORMATS.get(format, ("application/octet-stream", ".bin"))

    if export_type == "report" and report_html:
        filename = f"bluecircle_report_{datetime.now().strftime('%Y%m%d')}{ext}"
        if format == "pdf":
            buf = await export_report_pdf(report_html)
            return buf.getvalue(), content_type, filename
        elif format == "csv":
            content = report_html  # HTML as-is for CSV? Better to use the actual report CSV
            return content.encode("utf-8"), "text/html", filename.replace(ext, ".html")
        else:
            content = report_html
            return content.encode("utf-8"), "text/html", filename.replace(ext, ".html")

    if export_type == "tracks":
        if format == "csv":
            content = await export_tracks_csv(db, turtle_id=turtle_id)
            filename = f"bluecircle_tracks_{datetime.now().strftime('%Y%m%d')}{ext}"
            return content.encode("utf-8"), content_type, filename
        elif format == "excel":
            buf = await export_tracks_excel(db, turtle_id=turtle_id)
            filename = f"bluecircle_tracks_{datetime.now().strftime('%Y%m%d')}{ext}"
            return buf.getvalue(), content_type, filename

    elif export_type == "turtles":
        if format == "csv":
            content = await export_turtles_csv(db)
            filename = f"bluecircle_turtles_{datetime.now().strftime('%Y%m%d')}{ext}"
            return content.encode("utf-8"), content_type, filename

    elif export_type == "alerts":
        if format == "csv":
            content = await export_alerts_csv(db)
            filename = f"bluecircle_alerts_{datetime.now().strftime('%Y%m%d')}{ext}"
            return content.encode("utf-8"), content_type, filename

    raise ValueError(f"Unsupported export: type={export_type}, format={format}")
